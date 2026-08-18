"""Processamento leve da aba PROGRAMADO para o painel KM x Dia."""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook

COL_EMPRESA = 3
COL_FROTA = (11, 12, 13)
COL_VIAGENS = (14, 15, 16)
COL_KM_OPERACIONAL = 17
COL_KM_MORTA = 18
COL_DIA_INICIAL = 31
COL_DIA_FINAL = 61


def number(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                pass
    return None


def transporta(row) -> bool:
    return any("transporta" in str(row[index - 1] or "").casefold()
               for index in range(min(COL_KM_MORTA, len(row))))


def dates_in_period(ws, start_day: int, end_day: int):
    header = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), ())
    dates = []
    for column in range(COL_DIA_INICIAL, COL_DIA_FINAL + 1):
        value = to_date(header[column - 1] if len(header) >= column else None)
        if value and start_day <= value.day <= end_day:
            dates.append((column, value))
    return dates


def calculate(ws, start_day: int, end_day: int):
    dates = dates_in_period(ws, start_day, end_day)
    company_totals = defaultdict(lambda: {"km_operacional": 0.0, "km_morta": 0.0, "km_transporta": 0.0, "km_total": 0.0, "viagens": 0.0})
    daily = defaultdict(lambda: {"frota": 0.0, "viagens": 0.0, "km_operacional": 0.0, "km_morta": 0.0, "km_transporta": 0.0, "km_total": 0.0})
    company_daily = defaultdict(lambda: defaultdict(lambda: {"frota": 0.0, "viagens": 0.0, "km_operacional": 0.0, "km_morta": 0.0, "km_transporta": 0.0, "km_total": 0.0}))

    for row in ws.iter_rows(min_row=3, values_only=True):
        company = str(row[COL_EMPRESA - 1] if len(row) >= COL_EMPRESA else "").strip()
        if not company:
            continue
        op = number(row[COL_KM_OPERACIONAL - 1] if len(row) >= COL_KM_OPERACIONAL else 0)
        dead = number(row[COL_KM_MORTA - 1] if len(row) >= COL_KM_MORTA else 0)
        is_transporta = transporta(row)
        company_totals[company]["km_operacional"] += op
        company_totals[company]["km_morta"] += dead
        company_totals[company]["km_transporta"] += op if is_transporta else 0
        company_totals[company]["km_total"] += op + dead

        for column, current_date in dates:
            value = number(row[column - 1] if len(row) >= column else 0)
            if not value:
                continue
            key = current_date.isoformat()
            fleet = sum(number(row[c - 1] if len(row) >= c else 0) for c in COL_FROTA)
            trips = sum(number(row[c - 1] if len(row) >= c else 0) for c in COL_VIAGENS)
            for target in (daily[key], company_daily[company][key]):
                target["frota"] += fleet
                target["viagens"] += trips
                target["km_operacional"] += value
                target["km_transporta"] += value if is_transporta else 0
                target["km_total"] += value
    companies = [{"empresa": name, **values} for name, values in sorted(company_totals.items())]
    daily_rows = [{"data": key, **values} for key, values in sorted(daily.items())]
    company_daily_rows = {
        name: [{"data": key, **values} for key, values in sorted(values.items())]
        for name, values in company_daily.items()
    }
    return companies, daily_rows, company_daily_rows


def download_source(sheet_id: str) -> bytes:
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    response = requests.get(url, timeout=90)
    response.raise_for_status()
    return response.content


def build_report_from_bytes(source: bytes, destination: Path, start_day: int = 1, end_day: int = 31):
    workbook = load_workbook(io.BytesIO(source), data_only=True, read_only=True)
    if "PROGRAMADO" not in workbook.sheetnames:
        raise ValueError("A planilha precisa ter a aba PROGRAMADO.")
    companies, daily, company_daily = calculate(workbook["PROGRAMADO"], start_day, end_day)
    output = Workbook()
    output.remove(output.active)
    summary = output.create_sheet("TOTAL POR EMPRESA")
    summary.append(["Empresa", "KM Operacional", "KM Morta", "KM Transporta", "KM Total", "Viagens"])
    for company in companies:
        summary.append([company["empresa"], company["km_operacional"], company["km_morta"], company["km_transporta"], company["km_total"], company["viagens"]])
    destination.parent.mkdir(parents=True, exist_ok=True)
    output.save(destination)
    return {"companies": companies, "daily": daily, "company_daily": company_daily}


def build_report(sheet_id: str, destination: Path, start_day: int = 1, end_day: int = 31):
    return build_report_from_bytes(download_source(sheet_id), destination, start_day, end_day)
